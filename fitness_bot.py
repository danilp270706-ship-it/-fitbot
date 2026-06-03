import asyncio
import random
import io
from datetime import datetime, date, timedelta
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, MessageHandler, CallbackQueryHandler,
    filters, ContextTypes, ConversationHandler
)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from PIL import Image, ImageDraw, ImageFont
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

BOT_TOKEN = "8906586936:AAGLtPaukyp2PG22dTI17G7lxEZAJVCulH4"

(SETUP_NAME, SETUP_WEIGHT, SETUP_HEIGHT, SETUP_AGE, SETUP_GOAL) = range(5)

users = {}

# ── Мотивация ───────────────────────────────────────────────────────────────
MOTIVATIONAL_QUOTES = [
    "💪 Каждая тренировка делает тебя сильнее!",
    "🔥 Боль временна. Гордость — навсегда.",
    "🚀 Ты ближе к цели, чем вчера.",
    "⚡ Успех — это сумма маленьких усилий каждый день.",
    "🏆 Чемпионы делают то, что не хотят делать обычные люди.",
    "🌟 Прогресс, а не совершенство.",
    "🦁 Будь сильнее своих отмазок.",
    "✨ Сегодняшние усилия — завтрашние результаты.",
]

# ── База продуктов (100+ позиций) ──────────────────────────────────────────
FOOD_DB = {
    # Мясо и птица
    "куриная грудка": {"calories": 165, "protein": 31, "fat": 4, "carbs": 0},
    "куриное бедро": {"calories": 209, "protein": 26, "fat": 11, "carbs": 0},
    "говядина": {"calories": 250, "protein": 26, "fat": 17, "carbs": 0},
    "свинина": {"calories": 290, "protein": 25, "fat": 21, "carbs": 0},
    "индейка": {"calories": 189, "protein": 29, "fat": 7, "carbs": 0},
    "тунец": {"calories": 116, "protein": 26, "fat": 1, "carbs": 0},
    "лосось": {"calories": 208, "protein": 20, "fat": 13, "carbs": 0},
    "яйцо": {"calories": 78, "protein": 6, "fat": 5, "carbs": 1},
    "яичный белок": {"calories": 17, "protein": 4, "fat": 0, "carbs": 0},
    # Молочные
    "творог": {"calories": 101, "protein": 18, "fat": 2, "carbs": 4},
    "творог 0%": {"calories": 70, "protein": 16, "fat": 0, "carbs": 2},
    "молоко": {"calories": 62, "protein": 3, "fat": 3, "carbs": 5},
    "кефир": {"calories": 51, "protein": 3, "fat": 1, "carbs": 4},
    "греческий йогурт": {"calories": 97, "protein": 9, "fat": 5, "carbs": 4},
    "сыр": {"calories": 402, "protein": 25, "fat": 33, "carbs": 1},
    "протеин": {"calories": 120, "protein": 25, "fat": 1, "carbs": 3},
    # Крупы и злаки
    "овсянка": {"calories": 370, "protein": 13, "fat": 7, "carbs": 65},
    "рис": {"calories": 130, "protein": 3, "fat": 0, "carbs": 28},
    "гречка": {"calories": 343, "protein": 13, "fat": 3, "carbs": 68},
    "макароны": {"calories": 158, "protein": 6, "fat": 1, "carbs": 31},
    "хлеб белый": {"calories": 265, "protein": 9, "fat": 3, "carbs": 51},
    "хлеб ржаной": {"calories": 259, "protein": 9, "fat": 3, "carbs": 48},
    "картофель": {"calories": 77, "protein": 2, "fat": 0, "carbs": 17},
    "батат": {"calories": 86, "protein": 2, "fat": 0, "carbs": 20},
    # Фрукты
    "банан": {"calories": 89, "protein": 1, "fat": 0, "carbs": 23},
    "яблоко": {"calories": 52, "protein": 0, "fat": 0, "carbs": 14},
    "апельсин": {"calories": 47, "protein": 1, "fat": 0, "carbs": 12},
    "виноград": {"calories": 69, "protein": 1, "fat": 0, "carbs": 18},
    "клубника": {"calories": 32, "protein": 1, "fat": 0, "carbs": 8},
    "арбуз": {"calories": 30, "protein": 1, "fat": 0, "carbs": 7},
    # Овощи
    "огурец": {"calories": 15, "protein": 1, "fat": 0, "carbs": 3},
    "помидор": {"calories": 18, "protein": 1, "fat": 0, "carbs": 4},
    "морковь": {"calories": 41, "protein": 1, "fat": 0, "carbs": 10},
    "капуста": {"calories": 25, "protein": 1, "fat": 0, "carbs": 5},
    "брокколи": {"calories": 34, "protein": 3, "fat": 0, "carbs": 7},
    "шпинат": {"calories": 23, "protein": 3, "fat": 0, "carbs": 4},
    # Орехи и масла
    "арахис": {"calories": 567, "protein": 26, "fat": 49, "carbs": 16},
    "арахисовое масло": {"calories": 588, "protein": 25, "fat": 50, "carbs": 20},
    "миндаль": {"calories": 579, "protein": 21, "fat": 50, "carbs": 22},
    "грецкий орех": {"calories": 654, "protein": 15, "fat": 65, "carbs": 14},
    "оливковое масло": {"calories": 884, "protein": 0, "fat": 100, "carbs": 0},
    # Напитки
    "кофе": {"calories": 5, "protein": 0, "fat": 0, "carbs": 1},
    "молоко 2.5%": {"calories": 52, "protein": 3, "fat": 3, "carbs": 5},
    # Готовые блюда
    "пицца": {"calories": 266, "protein": 11, "fat": 10, "carbs": 33},
    "бургер": {"calories": 295, "protein": 17, "fat": 14, "carbs": 24},
    "борщ": {"calories": 50, "protein": 3, "fat": 2, "carbs": 6},
    "гречка с курицей": {"calories": 145, "protein": 17, "fat": 3, "carbs": 14},
}

# ── Программы тренировок ───────────────────────────────────────────────────
WORKOUT_PROGRAMS = {
    "fullbody": {
        "name": "💪 Full Body (3 дня)",
        "description": "Тренировки всего тела 3 раза в неделю",
        "days": {
            1: [("Приседания", 4, 10), ("Жим лёжа", 4, 10), ("Тяга в наклоне", 4, 10), ("Армейский жим", 3, 12), ("Планка", 3, 60)],
            2: [("Становая тяга", 4, 8), ("Жим гантелей", 4, 10), ("Подтягивания", 4, 8), ("Выпады", 3, 12), ("Пресс", 3, 20)],
            3: [("Фронтальные приседания", 4, 10), ("Жим на наклонной", 4, 10), ("Тяга верхнего блока", 4, 12), ("Разводка гантелей", 3, 15), ("Гиперэкстензия", 3, 15)],
        }
    },
    "split": {
        "name": "🏋️ Сплит (4 дня)",
        "description": "Грудь/трицепс, Спина/бицепс, Ноги, Плечи",
        "days": {
            1: [("Жим лёжа", 4, 10), ("Жим на наклонной", 4, 10), ("Разводка", 3, 15), ("Французский жим", 3, 12), ("Трицепс на блоке", 3, 15)],
            2: [("Становая тяга", 4, 8), ("Тяга в наклоне", 4, 10), ("Подтягивания", 4, 8), ("Тяга гантели", 3, 12), ("Подъём на бицепс", 4, 12)],
            3: [("Приседания", 4, 10), ("Жим ногами", 4, 12), ("Выпады", 3, 12), ("Разгибания ног", 3, 15), ("Сгибания ног", 3, 15)],
            4: [("Армейский жим", 4, 10), ("Тяга к подбородку", 4, 12), ("Махи гантелями", 3, 15), ("Шраги", 3, 15), ("Обратные махи", 3, 15)],
        }
    },
    "home": {
        "name": "🏠 Дома без железа (3 дня)",
        "description": "Только собственный вес",
        "days": {
            1: [("Отжимания", 4, 20), ("Приседания", 4, 25), ("Подъём ног лёжа", 3, 20), ("Планка", 3, 60), ("Берпи", 3, 10)],
            2: [("Широкие отжимания", 4, 15), ("Обратные отжимания", 4, 15), ("Выпады", 4, 20), ("Ягодичный мост", 4, 20), ("Скручивания", 3, 25)],
            3: [("Алмазные отжимания", 4, 12), ("Пистолет", 3, 8), ("Отжимания пайк", 3, 15), ("Прыжки", 3, 30), ("Велосипед", 3, 20)],
        }
    },
}

QUICK_FOODS = [
    {"name": "Яйцо варёное", "calories": 78, "protein": 6, "fat": 5, "carbs": 1},
    {"name": "Куриная грудка 100г", "calories": 165, "protein": 31, "fat": 4, "carbs": 0},
    {"name": "Овсянка 100г", "calories": 370, "protein": 13, "fat": 7, "carbs": 65},
    {"name": "Банан", "calories": 89, "protein": 1, "fat": 0, "carbs": 23},
    {"name": "Творог 100г", "calories": 101, "protein": 18, "fat": 2, "carbs": 4},
    {"name": "Кофе", "calories": 5, "protein": 0, "fat": 0, "carbs": 1},
    {"name": "Рис варёный 100г", "calories": 130, "protein": 3, "fat": 0, "carbs": 28},
    {"name": "Протеин", "calories": 120, "protein": 25, "fat": 1, "carbs": 3},
]

# ── Уровни ─────────────────────────────────────────────────────────────────
LEVELS = [
    (0,    "🌱 Новичок"),
    (500,  "💪 Любитель"),
    (1500, "⚡ Атлет"),
    (3000, "🏆 Профи"),
    (6000, "👑 Чемпион"),
    (10000,"🦁 Легенда"),
]

ACHIEVEMENTS = {
    "first_workout":   {"name": "🏋️ Первая тренировка",    "desc": "Записал первое упражнение"},
    "first_food":      {"name": "🍽 Первый приём пищи",     "desc": "Записал первую еду"},
    "water_day":       {"name": "💧 Водный день",           "desc": "Выполнил норму воды за день"},
    "calorie_goal":    {"name": "🎯 В цель!",               "desc": "Попал в норму калорий"},
    "streak_3":        {"name": "🔥 3 дня подряд",          "desc": "Активен 3 дня подряд"},
    "streak_7":        {"name": "🔥🔥 Неделя!",             "desc": "Активен 7 дней подряд"},
    "streak_30":       {"name": "🔥🔥🔥 Месяц!",            "desc": "Активен 30 дней подряд"},
    "record_broken":   {"name": "💥 Личный рекорд!",        "desc": "Побил личный рекорд в упражнении"},
    "workouts_10":     {"name": "🏅 10 тренировок",         "desc": "Провёл 10 тренировок"},
    "workouts_50":     {"name": "🥇 50 тренировок",         "desc": "Провёл 50 тренировок"},
    "weight_lost_5":   {"name": "📉 −5 кг",                 "desc": "Похудел на 5 кг"},
}

# ── Хелперы ────────────────────────────────────────────────────────────────
def get_user(uid):
    if uid not in users:
        users[uid] = {
            "name": None, "weight": None, "height": None, "age": None,
            "goal": None, "calorie_goal": 2000, "water_goal": 2000,
            "today": str(date.today()),
            "food_log": [], "workout_log": [], "water": 0,
            "reminders": {"water": False, "food": False},
            "streak": 0, "last_active": None,
            "weight_history": [],
            "program": None, "program_day": 1,
            "history": {},
            "xp": 0,
            "achievements": [],
            "records": {},
            "total_workouts": 0,
        }
    u = users[uid]
    today = str(date.today())
    if u["today"] != today:
        total_cal, p, f, c = calories_summary(u)
        if u["food_log"] or u["workout_log"]:
            u["history"][u["today"]] = {
                "calories": total_cal, "protein": p, "fat": f, "carbs": c,
                "water": u["water"], "workouts": len(u["workout_log"]),
                "food_log": u["food_log"][:], "workout_log": u["workout_log"][:],
            }
        yesterday = str(date.today() - timedelta(days=1))
        if u["last_active"] == yesterday:
            u["streak"] += 1
        elif u["last_active"] != yesterday and u["last_active"] is not None:
            u["streak"] = 0
        u["last_active"] = today
        u["today"] = today
        u["food_log"] = []
        u["workout_log"] = []
        u["water"] = 0
    return u

def calories_summary(u):
    return (
        sum(f["calories"] for f in u["food_log"]),
        sum(f["protein"] for f in u["food_log"]),
        sum(f["fat"] for f in u["food_log"]),
        sum(f["carbs"] for f in u["food_log"]),
    )

def progress_bar(value, maximum, length=10):
    filled = min(int(value / maximum * length), length) if maximum > 0 else 0
    return "▓" * filled + "░" * (length - filled)

def get_level(xp):
    level_name = LEVELS[0][1]
    for threshold, name in LEVELS:
        if xp >= threshold:
            level_name = name
    return level_name

def get_next_level(xp):
    for i, (threshold, name) in enumerate(LEVELS):
        if xp < threshold:
            return threshold, name
    return None, None

def add_xp(u, amount):
    u["xp"] = u.get("xp", 0) + amount

def check_achievements(u):
    new = []
    earned = set(u.get("achievements", []))

    if u["workout_log"] and "first_workout" not in earned:
        new.append("first_workout"); earned.add("first_workout")
    if u["food_log"] and "first_food" not in earned:
        new.append("first_food"); earned.add("first_food")
    if u["water"] >= u["water_goal"] and "water_day" not in earned:
        new.append("water_day"); earned.add("water_day")
    total_cal, _, _, _ = calories_summary(u)
    if abs(total_cal - u["calorie_goal"]) <= u["calorie_goal"] * 0.1 and total_cal > 0 and "calorie_goal" not in earned:
        new.append("calorie_goal"); earned.add("calorie_goal")
    if u["streak"] >= 3 and "streak_3" not in earned:
        new.append("streak_3"); earned.add("streak_3")
    if u["streak"] >= 7 and "streak_7" not in earned:
        new.append("streak_7"); earned.add("streak_7")
    if u["streak"] >= 30 and "streak_30" not in earned:
        new.append("streak_30"); earned.add("streak_30")
    if u.get("total_workouts", 0) >= 10 and "workouts_10" not in earned:
        new.append("workouts_10"); earned.add("workouts_10")
    if u.get("total_workouts", 0) >= 50 and "workouts_50" not in earned:
        new.append("workouts_50"); earned.add("workouts_50")

    u["achievements"] = list(earned)
    return new

def search_food(query):
    query = query.lower().strip()
    results = []
    for name, data in FOOD_DB.items():
        if query in name:
            results.append((name, data))
    return results[:6]

def ai_advice(u):
    total_cal, p, f, c = calories_summary(u)
    goal = u["calorie_goal"]
    tips = []
    weight = u["weight"] or 70
    rec_protein = weight * 1.6
    if p < rec_protein * 0.7:
        tips.append(f"🥩 Маловато белка ({p}г из ~{int(rec_protein)}г). Добавь куриную грудку или творог.")
    else:
        tips.append(f"✅ Белок в норме! ({p}г)")
    if total_cal > goal * 1.1:
        tips.append(f"⚠️ Превышение на {total_cal - goal} ккал. Снизь ужин.")
    elif total_cal < goal * 0.5 and datetime.now().hour > 14:
        tips.append(f"📉 Съедено только {total_cal}/{goal} ккал. Не голодай!")
    else:
        tips.append(f"✅ Калории в норме! {total_cal}/{goal} ккал.")
    if u["water"] < u["water_goal"] * 0.5 and datetime.now().hour > 12:
        tips.append(f"💧 Только {u['water']}мл воды. Обезвоживание снижает силу на 20%!")
    else:
        tips.append(f"💧 Вода в норме! {u['water']}мл.")
    return "\n\n".join(tips)

def make_weekly_chart(u):
    if not PIL_AVAILABLE:
        return None
    days = []
    cals = []
    for i in range(6, -1, -1):
        d = str(date.today() - timedelta(days=i))
        label = (date.today() - timedelta(days=i)).strftime("%d.%m")
        if d == str(date.today()):
            total_cal, _, _, _ = calories_summary(u)
            cals.append(total_cal)
        else:
            cals.append(u.get("history", {}).get(d, {}).get("calories", 0))
        days.append(label)

    W, H = 700, 400
    img = Image.new("RGB", (W, H), (18, 18, 18))
    draw = ImageDraw.Draw(img)

    try:
        font = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", 22)
        font_small = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", 16)
        font_title = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", 28)
    except:
        font = ImageFont.load_default()
        font_small = font
        font_title = font

    draw.text((20, 15), "📊 Калории за 7 дней", font=font_title, fill=(255, 255, 255))

    pad_l, pad_r, pad_t, pad_b = 60, 30, 60, 60
    chart_w = W - pad_l - pad_r
    chart_h = H - pad_t - pad_b

    max_cal = max(max(cals), u["calorie_goal"]) * 1.1 or 2000

    # Сетка
    for i in range(5):
        y = pad_t + chart_h - int(chart_h * i / 4)
        val = int(max_cal * i / 4)
        draw.line([(pad_l, y), (W - pad_r, y)], fill=(40, 40, 40), width=1)
        draw.text((5, y - 10), str(val), font=font_small, fill=(120, 120, 120))

    # Линия нормы калорий
    goal_y = pad_t + chart_h - int(chart_h * u["calorie_goal"] / max_cal)
    draw.line([(pad_l, goal_y), (W - pad_r, goal_y)], fill=(255, 200, 0), width=2)
    draw.text((W - pad_r - 60, goal_y - 20), "Норма", font=font_small, fill=(255, 200, 0))

    bar_w = chart_w // len(days)
    for i, (day, cal) in enumerate(zip(days, cals)):
        x = pad_l + i * bar_w + bar_w // 6
        bar_h = int(chart_h * cal / max_cal) if max_cal > 0 else 0
        y1 = pad_t + chart_h - bar_h
        y2 = pad_t + chart_h
        color = (0, 200, 100) if cal <= u["calorie_goal"] else (220, 80, 80)
        if cal == 0:
            color = (60, 60, 60)
        draw.rounded_rectangle([x, y1, x + bar_w * 2 // 3, y2], radius=6, fill=color)
        if cal > 0:
            draw.text((x, y1 - 22), str(cal), font=font_small, fill=(220, 220, 220))
        draw.text((x - 5, H - pad_b + 8), day, font=font_small, fill=(180, 180, 180))

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return buf

def main_menu(u):
    total_cal, p, f, c = calories_summary(u)
    goal = u["calorie_goal"]
    water = u["water"]
    water_goal = u["water_goal"]
    streak = u["streak"]
    level = get_level(u.get("xp", 0))
    xp = u.get("xp", 0)
    next_thresh, next_level = get_next_level(xp)
    xp_text = f" | ⭐ {xp} XP" if not next_thresh else f" | ⭐ {xp}/{next_thresh} XP"

    text = (
        f"{level}{xp_text}\n"
        f"👋 *{u['name'] or 'Атлет'}*"
        + (f"  🔥 Стрик: *{streak} дн.*" if streak > 0 else "") + "\n"
        f"📅 {date.today().strftime('%d.%m.%Y')}\n\n"
        f"🔥 Калории: *{total_cal}* / {goal} ккал\n"
        f"{progress_bar(total_cal, goal)} {int(total_cal/goal*100) if goal else 0}%\n\n"
        f"🥩 Б: *{p}г*  🧈 Ж: *{f}г*  🍞 У: *{c}г*\n\n"
        f"💧 Вода: *{water}мл* / {water_goal}мл\n"
        f"{progress_bar(water, water_goal)} {int(water/water_goal*100) if water_goal else 0}%\n\n"
        f"🏋️ Тренировок: *{len(u['workout_log'])}*  |  Осталось: *{max(goal-total_cal,0)} ккал*"
    )
    keyboard = [
        [InlineKeyboardButton("🍽 Добавить еду", callback_data="add_food"),
         InlineKeyboardButton("🔍 Поиск еды", callback_data="search_food")],
        [InlineKeyboardButton("⚡ Быстрая еда", callback_data="quick_food"),
         InlineKeyboardButton("💧 +250мл воды", callback_data="add_water_250")],
        [InlineKeyboardButton("🏋️ Тренировка", callback_data="add_workout"),
         InlineKeyboardButton("📋 Программа", callback_data="programs")],
        [InlineKeyboardButton("🏆 Рекорды", callback_data="records"),
         InlineKeyboardButton("🎖 Достижения", callback_data="achievements")],
        [InlineKeyboardButton("🤖 ИИ-совет", callback_data="ai_advice"),
         InlineKeyboardButton("📊 Статистика", callback_data="stats")],
        [InlineKeyboardButton("📈 Прогресс веса", callback_data="weight_progress"),
         InlineKeyboardButton("📤 Экспорт Excel", callback_data="export_excel")],
        [InlineKeyboardButton("⚙️ Настройки", callback_data="settings")],
    ]
    return text, InlineKeyboardMarkup(keyboard)

# ── Настройка профиля ──────────────────────────────────────────────────────
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    u = get_user(uid)
    if not u["name"]:
        await update.message.reply_text(
            "💪 *Добро пожаловать в FitBot Pro!*\n\nУчёт питания, тренировки, уровни, достижения и ИИ-советник.\n\nКак тебя зовут?",
            parse_mode="Markdown"
        )
        return SETUP_NAME
    text, markup = main_menu(u)
    await update.message.reply_text(text, reply_markup=markup, parse_mode="Markdown")
    return ConversationHandler.END

async def setup_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    u = get_user(uid)
    u["name"] = update.message.text.strip()
    await update.message.reply_text(f"Отлично, *{u['name']}*! 💪\n\nСколько ты весишь? (кг)", parse_mode="Markdown")
    return SETUP_WEIGHT

async def setup_weight(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    u = get_user(uid)
    try:
        u["weight"] = float(update.message.text.replace(",", "."))
        u["weight_history"].append({"date": str(date.today()), "weight": u["weight"]})
        await update.message.reply_text("Твой рост? (см)")
        return SETUP_HEIGHT
    except:
        await update.message.reply_text("Введи число:")
        return SETUP_WEIGHT

async def setup_height(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    u = get_user(uid)
    try:
        u["height"] = float(update.message.text.replace(",", "."))
        await update.message.reply_text("Сколько тебе лет?")
        return SETUP_AGE
    except:
        await update.message.reply_text("Введи число:")
        return SETUP_HEIGHT

async def setup_age(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    u = get_user(uid)
    try:
        u["age"] = int(update.message.text)
        keyboard = [
            [InlineKeyboardButton("🏃 Похудеть", callback_data="goal_lose"),
             InlineKeyboardButton("💪 Набрать массу", callback_data="goal_gain")],
            [InlineKeyboardButton("⚖️ Поддержать форму", callback_data="goal_maintain")],
        ]
        await update.message.reply_text("Какая у тебя цель?", reply_markup=InlineKeyboardMarkup(keyboard))
        return SETUP_GOAL
    except:
        await update.message.reply_text("Введи число:")
        return SETUP_AGE

async def setup_goal(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    uid = query.from_user.id
    u = get_user(uid)
    goals = {"goal_lose": "похудеть", "goal_gain": "набрать массу", "goal_maintain": "поддержать форму"}
    calories = {"goal_lose": 1700, "goal_gain": 2800, "goal_maintain": 2200}
    u["goal"] = goals[query.data]
    u["calorie_goal"] = calories[query.data]
    add_xp(u, 100)
    text, markup = main_menu(u)
    await query.edit_message_text(f"✅ Цель: *{u['goal']}* | Норма: *{u['calorie_goal']} ккал/день*\n\n+100 XP за регистрацию! 🎉", parse_mode="Markdown")
    await query.message.reply_text(text, reply_markup=markup, parse_mode="Markdown")
    return ConversationHandler.END

async def menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    u = get_user(uid)
    if not u["name"]:
        await update.message.reply_text("Сначала /start")
        return
    text, markup = main_menu(u)
    await update.message.reply_text(text, reply_markup=markup, parse_mode="Markdown")

# ── Кнопки ──────────────────────────────────────────────────────────────────
async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    uid = query.from_user.id
    u = get_user(uid)
    data = query.data

    # Вода
    if data == "add_water_250":
        u["water"] += 250
        add_xp(u, 5)
        text, markup = main_menu(u)
        await query.edit_message_text(f"💧 +250мл! Итого: {u['water']}мл\n\n{text}", reply_markup=markup, parse_mode="Markdown")

    # Поиск еды
    elif data == "search_food":
        context.user_data["state"] = "search_food"
        keyboard = [[InlineKeyboardButton("❌ Отмена", callback_data="back_menu")]]
        await query.edit_message_text(
            "🔍 *Поиск продукта*\n\nНапиши название (например: курица, рис, банан):",
            reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="Markdown"
        )

    # Быстрая еда
    elif data == "quick_food":
        keyboard = []
        row = []
        for i, food in enumerate(QUICK_FOODS):
            row.append(InlineKeyboardButton(food["name"], callback_data=f"qf_{i}"))
            if len(row) == 2:
                keyboard.append(row)
                row = []
        if row:
            keyboard.append(row)
        keyboard.append([InlineKeyboardButton("⬅️ Назад", callback_data="back_menu")])
        await query.edit_message_text("⚡ *Быстрое добавление:*", reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="Markdown")

    elif data.startswith("qf_"):
        idx = int(data.replace("qf_", ""))
        food = QUICK_FOODS[idx]
        u["food_log"].append({**food, "time": datetime.now().strftime("%H:%M")})
        add_xp(u, 10)
        new_ach = check_achievements(u)
        text, markup = main_menu(u)
        msg = f"✅ *{food['name']}* добавлено! ({food['calories']} ккал)\n\n{text}"
        if new_ach:
            msg = "🎖 *Новое достижение: " + ACHIEVEMENTS[new_ach[0]]["name"] + "!*\n\n" + msg
        await query.edit_message_text(msg, reply_markup=markup, parse_mode="Markdown")

    # Результаты поиска еды
    elif data.startswith("fd_"):
        food_name = data[3:]
        if food_name in FOOD_DB:
            food = FOOD_DB[food_name]
            context.user_data["state"] = "food_grams"
            context.user_data["food_db_name"] = food_name
            context.user_data["food_db_data"] = food
            keyboard = [
                [InlineKeyboardButton("100г", callback_data="fg_100"),
                 InlineKeyboardButton("150г", callback_data="fg_150"),
                 InlineKeyboardButton("200г", callback_data="fg_200")],
                [InlineKeyboardButton("250г", callback_data="fg_250"),
                 InlineKeyboardButton("300г", callback_data="fg_300"),
                 InlineKeyboardButton("✏️ Своя порция", callback_data="fg_custom")],
                [InlineKeyboardButton("❌ Отмена", callback_data="back_menu")],
            ]
            f = food
            await query.edit_message_text(
                f"🍽 *{food_name.capitalize()}* (на 100г):\n"
                f"Калории: {f['calories']} | Б: {f['protein']}г | Ж: {f['fat']}г | У: {f['carbs']}г\n\n"
                f"Выбери порцию:",
                reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="Markdown"
            )

    elif data.startswith("fg_"):
        portion = data[3:]
        food_name = context.user_data.get("food_db_name")
        food_base = context.user_data.get("food_db_data")
        if portion == "custom":
            context.user_data["state"] = "food_grams_input"
            keyboard = [[InlineKeyboardButton("❌ Отмена", callback_data="back_menu")]]
            await query.edit_message_text("Введи вес порции в граммах:", reply_markup=InlineKeyboardMarkup(keyboard))
        else:
            grams = int(portion)
            ratio = grams / 100
            entry = {
                "name": f"{food_name.capitalize()} {grams}г",
                "calories": round(food_base["calories"] * ratio),
                "protein": round(food_base["protein"] * ratio),
                "fat": round(food_base["fat"] * ratio),
                "carbs": round(food_base["carbs"] * ratio),
                "time": datetime.now().strftime("%H:%M"),
            }
            u["food_log"].append(entry)
            add_xp(u, 10)
            new_ach = check_achievements(u)
            context.user_data.pop("state", None)
            t, m = main_menu(u)
            msg = f"✅ *{entry['name']}* добавлено! ({entry['calories']} ккал)\n\n{t}"
            if new_ach:
                msg = "🎖 *" + ACHIEVEMENTS[new_ach[0]]["name"] + "!*\n\n" + msg
            await query.edit_message_text(msg, reply_markup=m, parse_mode="Markdown")

    # Ручное добавление еды
    elif data == "add_food":
        context.user_data["state"] = "food_name"
        keyboard = [[InlineKeyboardButton("❌ Отмена", callback_data="back_menu")]]
        await query.edit_message_text("🍽 *Добавить еду*\n\nНапиши название:", reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="Markdown")

    # Тренировки
    elif data == "add_workout":
        keyboard = [
            [InlineKeyboardButton("🏋️ Жим лёжа", callback_data="workout_Жим лёжа"),
             InlineKeyboardButton("🦵 Приседания", callback_data="workout_Приседания")],
            [InlineKeyboardButton("🔝 Становая тяга", callback_data="workout_Становая тяга"),
             InlineKeyboardButton("💪 Подтягивания", callback_data="workout_Подтягивания")],
            [InlineKeyboardButton("🏃 Бег", callback_data="workout_Бег"),
             InlineKeyboardButton("✏️ Своё", callback_data="workout_custom")],
            [InlineKeyboardButton("⬅️ Назад", callback_data="back_menu")],
        ]
        await query.edit_message_text("🏋️ *Выбери упражнение:*", reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="Markdown")

    elif data.startswith("workout_") and data != "workout_custom":
        name = data.replace("workout_", "")
        context.user_data["workout_name"] = name
        context.user_data["state"] = "workout_sets"
        rec = u.get("records", {}).get(name)
        rec_text = f"\n🏆 Твой рекорд: {rec['weight']}кг × {rec['reps']} повт." if rec else ""
        keyboard = [[InlineKeyboardButton("❌ Отмена", callback_data="back_menu")]]
        await query.edit_message_text(f"🏋️ *{name}*{rec_text}\n\nСколько подходов?", reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="Markdown")

    elif data == "workout_custom":
        context.user_data["state"] = "workout_custom_name"
        keyboard = [[InlineKeyboardButton("❌ Отмена", callback_data="back_menu")]]
        await query.edit_message_text("✏️ Напиши название упражнения:", reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="Markdown")

    # Рекорды
    elif data == "records":
        records = u.get("records", {})
        if not records:
            text = "🏆 *Личные рекорды*\n\nПока нет рекордов. Запиши тренировку!"
        else:
            text = "🏆 *Личные рекорды:*\n\n"
            for ex, r in sorted(records.items()):
                text += f"• *{ex}*: {r['weight']}кг × {r['reps']} повт. [{r['date']}]\n"
        keyboard = [[InlineKeyboardButton("⬅️ Назад", callback_data="back_menu")]]
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="Markdown")

    # Достижения
    elif data == "achievements":
        earned = set(u.get("achievements", []))
        xp = u.get("xp", 0)
        level = get_level(xp)
        next_thresh, next_name = get_next_level(xp)
        text = f"🎖 *Достижения*\n\n{level} | ⭐ {xp} XP\n"
        if next_thresh:
            text += f"До *{next_name}*: {next_thresh - xp} XP\n"
        text += "\n"
        for key, ach in ACHIEVEMENTS.items():
            if key in earned:
                text += f"✅ {ach['name']}\n_{ach['desc']}_\n\n"
            else:
                text += f"🔒 {ach['name']}\n_{ach['desc']}_\n\n"
        keyboard = [[InlineKeyboardButton("⬅️ Назад", callback_data="back_menu")]]
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="Markdown")

    # Программы
    elif data == "programs":
        current = u.get("program")
        current_text = f"\n\n✅ Активна: *{WORKOUT_PROGRAMS[current]['name']}*, день {u['program_day']}" if current else ""
        keyboard = [
            [InlineKeyboardButton(WORKOUT_PROGRAMS["fullbody"]["name"], callback_data="prog_fullbody")],
            [InlineKeyboardButton(WORKOUT_PROGRAMS["split"]["name"], callback_data="prog_split")],
            [InlineKeyboardButton(WORKOUT_PROGRAMS["home"]["name"], callback_data="prog_home")],
        ]
        if current:
            keyboard.append([InlineKeyboardButton("▶️ Следующая тренировка", callback_data="prog_next")])
        keyboard.append([InlineKeyboardButton("⬅️ Назад", callback_data="back_menu")])
        await query.edit_message_text(f"📋 *Программы тренировок*{current_text}", reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="Markdown")

    elif data.startswith("prog_") and data != "prog_next":
        prog_key = data.replace("prog_", "")
        u["program"] = prog_key
        u["program_day"] = 1
        prog = WORKOUT_PROGRAMS[prog_key]
        keyboard = [
            [InlineKeyboardButton("▶️ Начать день 1", callback_data="prog_next")],
            [InlineKeyboardButton("⬅️ Назад", callback_data="programs")],
        ]
        await query.edit_message_text(f"✅ *{prog['name']}* выбрана!\n_{prog['description']}_", reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="Markdown")

    elif data == "prog_next":
        prog_key = u.get("program")
        if not prog_key:
            return
        prog = WORKOUT_PROGRAMS[prog_key]
        day = u["program_day"]
        exercises = prog["days"].get(day, [])
        text = f"📋 *{prog['name']} — День {day}*\n\n"
        for ex_name, sets, reps in exercises:
            text += f"• *{ex_name}*: {sets}×{reps}\n"
            u["workout_log"].append({"name": ex_name, "sets": sets, "reps": reps, "weight": 0, "time": datetime.now().strftime("%H:%M")})
        u["total_workouts"] = u.get("total_workouts", 0) + len(exercises)
        add_xp(u, 50)
        u["program_day"] = (day % len(prog["days"])) + 1
        new_ach = check_achievements(u)
        keyboard = [[InlineKeyboardButton("⬅️ В меню", callback_data="back_menu")]]
        ach_text = ("\n🎖 " + ACHIEVEMENTS[new_ach[0]]["name"] + "!") if new_ach else ""
        await query.edit_message_text(text + f"\n✅ Записано! +50 XP{ach_text}\nСледующий день: {u['program_day']}", reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="Markdown")

    # ИИ-совет
    elif data == "ai_advice":
        advice = ai_advice(u)
        keyboard = [[InlineKeyboardButton("⬅️ Назад", callback_data="back_menu")]]
        await query.edit_message_text(f"🤖 *ИИ-анализ:*\n\n{advice}", reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="Markdown")

    # Статистика + график
    elif data == "stats":
        total_cal, p, f, c = calories_summary(u)
        workout_text = "\n".join(f"• {w['name']}: {w['sets']}×{w['reps']}" + (f" @ {w['weight']}кг" if w['weight'] else "") for w in u["workout_log"]) or "Нет"
        food_text = "\n".join(f"• {f['name']}: {f['calories']} ккал" for f in u["food_log"]) or "Нет"
        week_cal = sum(u.get("history", {}).get(str(date.today() - timedelta(days=i)), {}).get("calories", 0) for i in range(7))
        week_w = sum(u.get("history", {}).get(str(date.today() - timedelta(days=i)), {}).get("workouts", 0) for i in range(7))
        text = (
            f"📊 *Сегодня — {date.today().strftime('%d.%m.%Y')}*\n\n"
            f"🍽 *Еда:*\n{food_text}\n*Итого:* {total_cal} ккал | Б:{p}г Ж:{f}г У:{c}г\n\n"
            f"🏋️ *Тренировки:*\n{workout_text}\n\n"
            f"💧 Вода: {u['water']}мл\n\n"
            f"📅 *За 7 дней:* ~{int(week_cal/7)} ккал/день, тренировок: {week_w}\n"
            f"🔥 Стрик: {u['streak']} дней | ⭐ {u.get('xp',0)} XP"
        )
        keyboard = [
            [InlineKeyboardButton("📈 График калорий", callback_data="weekly_chart")],
            [InlineKeyboardButton("⬅️ Назад", callback_data="back_menu")],
        ]
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="Markdown")

    elif data == "weekly_chart":
        chart = make_weekly_chart(u)
        if chart:
            await query.message.reply_photo(photo=chart, caption="📊 Калории за 7 дней")
        else:
            await query.message.reply_text("⚠️ Установи Pillow:\n`pip3 install Pillow`", parse_mode="Markdown")
        t, m = main_menu(u)
        await query.edit_message_text(t, reply_markup=m, parse_mode="Markdown")

    # Прогресс веса
    elif data == "weight_progress":
        history = u.get("weight_history", [])
        if len(history) < 2:
            text = f"📈 *Прогресс веса*\n\nТекущий вес: *{u['weight']} кг*\n\nОбновляй вес регулярно командой /weight"
        else:
            first, last = history[0], history[-1]
            diff = last["weight"] - first["weight"]
            sign = "+" if diff > 0 else ""
            text = "📈 *Прогресс веса*\n\n"
            for entry in history[-10:]:
                text += f"📅 {entry['date']}: *{entry['weight']} кг*\n"
            text += f"\n{'📉' if diff < 0 else '📈'} Изменение: *{sign}{diff:.1f} кг*"
        keyboard = [
            [InlineKeyboardButton("⚖️ Обновить вес", callback_data="update_weight")],
            [InlineKeyboardButton("⬅️ Назад", callback_data="back_menu")],
        ]
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="Markdown")

    elif data == "update_weight":
        context.user_data["state"] = "update_weight"
        keyboard = [[InlineKeyboardButton("❌ Отмена", callback_data="back_menu")]]
        await query.edit_message_text(f"⚖️ Текущий: *{u['weight']} кг*\n\nВведи новый:", reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="Markdown")

    # Экспорт Excel
    elif data == "export_excel":
        if not EXCEL_AVAILABLE:
            keyboard = [[InlineKeyboardButton("⬅️ Назад", callback_data="back_menu")]]
            await query.edit_message_text("⚠️ Установи:\n`pip3 install openpyxl`", reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="Markdown")
            return
        await query.edit_message_text("📤 Генерирую отчёт...")
        wb = openpyxl.Workbook()
        hf = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        hfont = Font(color="FFFFFF", bold=True)

        ws1 = wb.active
        ws1.title = "Питание"
        for col, h in enumerate(["Дата","Продукт","Калории","Белки","Жиры","Углеводы","Время"], 1):
            c = ws1.cell(row=1, column=col, value=h); c.font = hfont; c.fill = hf

        all_days = {**u.get("history", {})}
        if u["food_log"]:
            tc, p, f, c2 = calories_summary(u)
            all_days[str(date.today())] = {"food_log": u["food_log"][:], "workout_log": u["workout_log"][:], "calories": tc, "protein": p, "fat": f, "carbs": c2, "water": u["water"], "workouts": len(u["workout_log"])}

        row = 2
        for day, dd in sorted(all_days.items()):
            for food in dd.get("food_log", []):
                for col, val in enumerate([day, food.get("name",""), food.get("calories",0), food.get("protein",0), food.get("fat",0), food.get("carbs",0), food.get("time","")], 1):
                    ws1.cell(row=row, column=col, value=val)
                row += 1

        ws2 = wb.create_sheet("Тренировки")
        for col, h in enumerate(["Дата","Упражнение","Подходы","Повторения","Вес (кг)","Время"], 1):
            c = ws2.cell(row=1, column=col, value=h); c.font = hfont; c.fill = hf
        row = 2
        for day, dd in sorted(all_days.items()):
            for w in dd.get("workout_log", []):
                for col, val in enumerate([day, w.get("name",""), w.get("sets",0), w.get("reps",0), w.get("weight",0), w.get("time","")], 1):
                    ws2.cell(row=row, column=col, value=val)
                row += 1

        ws3 = wb.create_sheet("Сводка")
        for col, h in enumerate(["Дата","Калории","Белки","Жиры","Углеводы","Вода","Тренировки"], 1):
            c = ws3.cell(row=1, column=col, value=h); c.font = hfont; c.fill = hf
        for ri, (day, dd) in enumerate(sorted(all_days.items()), 2):
            for col, val in enumerate([day, dd.get("calories",0), dd.get("protein",0), dd.get("fat",0), dd.get("carbs",0), dd.get("water",0), dd.get("workouts",0)], 1):
                ws3.cell(row=ri, column=col, value=val)

        ws4 = wb.create_sheet("Рекорды")
        for col, h in enumerate(["Упражнение","Вес (кг)","Повторения","Дата"], 1):
            c = ws4.cell(row=1, column=col, value=h); c.font = hfont; c.fill = hf
        for ri, (ex, r) in enumerate(u.get("records", {}).items(), 2):
            ws4.cell(row=ri, column=1, value=ex); ws4.cell(row=ri, column=2, value=r["weight"])
            ws4.cell(row=ri, column=3, value=r["reps"]); ws4.cell(row=ri, column=4, value=r["date"])

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        await query.message.reply_document(document=buf, filename=f"fitbot_{u['name']}_{date.today()}.xlsx", caption=f"📊 Отчёт готов, *{u['name']}*!", parse_mode="Markdown")
        t, m = main_menu(u)
        await query.message.reply_text(t, reply_markup=m, parse_mode="Markdown")

    # Настройки
    elif data == "settings":
        r = u["reminders"]
        keyboard = [
            [InlineKeyboardButton(f"💧 Вода: {'✅' if r['water'] else '❌'}", callback_data="toggle_water")],
            [InlineKeyboardButton(f"🍽 Еда: {'✅' if r['food'] else '❌'}", callback_data="toggle_food")],
            [InlineKeyboardButton("🎯 Норма калорий", callback_data="set_cal_goal")],
            [InlineKeyboardButton("💧 Норма воды", callback_data="set_water_goal")],
            [InlineKeyboardButton("⬅️ Назад", callback_data="back_menu")],
        ]
        await query.edit_message_text("⚙️ *Настройки*", reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="Markdown")

    elif data in ("toggle_water", "toggle_food"):
        key = "water" if data == "toggle_water" else "food"
        u["reminders"][key] = not u["reminders"][key]
        r = u["reminders"]
        keyboard = [
            [InlineKeyboardButton(f"💧 Вода: {'✅' if r['water'] else '❌'}", callback_data="toggle_water")],
            [InlineKeyboardButton(f"🍽 Еда: {'✅' if r['food'] else '❌'}", callback_data="toggle_food")],
            [InlineKeyboardButton("🎯 Норма калорий", callback_data="set_cal_goal")],
            [InlineKeyboardButton("💧 Норма воды", callback_data="set_water_goal")],
            [InlineKeyboardButton("⬅️ Назад", callback_data="back_menu")],
        ]
        await query.edit_message_text("⚙️ *Настройки*", reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="Markdown")

    elif data == "set_cal_goal":
        context.user_data["state"] = "set_calorie_goal"
        keyboard = [[InlineKeyboardButton("❌ Отмена", callback_data="back_menu")]]
        await query.edit_message_text("🔥 Введи норму калорий:", reply_markup=InlineKeyboardMarkup(keyboard))

    elif data == "set_water_goal":
        context.user_data["state"] = "set_water_goal"
        keyboard = [[InlineKeyboardButton("❌ Отмена", callback_data="back_menu")]]
        await query.edit_message_text("💧 Введи норму воды в мл:", reply_markup=InlineKeyboardMarkup(keyboard))

    elif data == "back_menu":
        context.user_data.pop("state", None)
        text, markup = main_menu(u)
        await query.edit_message_text(text, reply_markup=markup, parse_mode="Markdown")

# ── Текст ──────────────────────────────────────────────────────────────────
async def text_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    u = get_user(uid)
    state = context.user_data.get("state")
    text = update.message.text.strip()

    # Поиск еды
    if state == "search_food":
        results = search_food(text)
        if not results:
            await update.message.reply_text(f"❌ «{text}» не найдено. Попробуй другое слово или добавь вручную.")
            return
        keyboard = []
        for name, data in results:
            label = f"{name.capitalize()} ({data['calories']} ккал)"
            keyboard.append([InlineKeyboardButton(label, callback_data=f"fd_{name}")])
        keyboard.append([InlineKeyboardButton("❌ Отмена", callback_data="back_menu")])
        await update.message.reply_text(
            f"🔍 Найдено по запросу «{text}»:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    elif state == "food_grams_input":
        try:
            grams = int(text)
            food_name = context.user_data.get("food_db_name")
            food_base = context.user_data.get("food_db_data")
            ratio = grams / 100
            entry = {
                "name": f"{food_name.capitalize()} {grams}г",
                "calories": round(food_base["calories"] * ratio),
                "protein": round(food_base["protein"] * ratio),
                "fat": round(food_base["fat"] * ratio),
                "carbs": round(food_base["carbs"] * ratio),
                "time": datetime.now().strftime("%H:%M"),
            }
            u["food_log"].append(entry)
            add_xp(u, 10)
            new_ach = check_achievements(u)
            context.user_data.pop("state", None)
            t, m = main_menu(u)
            msg = f"✅ *{entry['name']}* добавлено! ({entry['calories']} ккал)\n\n{t}"
            if new_ach:
                msg = "🎖 *" + ACHIEVEMENTS[new_ach[0]]["name"] + "!*\n\n" + msg
            await update.message.reply_text(msg, reply_markup=m, parse_mode="Markdown")
        except:
            await update.message.reply_text("Введи число:")

    elif state == "food_name":
        context.user_data["food_name"] = text
        context.user_data["state"] = "food_calories"
        await update.message.reply_text(f"🍽 *{text}*\n\nКалорий:", parse_mode="Markdown")

    elif state == "food_calories":
        try:
            context.user_data["food_calories"] = int(text)
            context.user_data["state"] = "food_protein"
            await update.message.reply_text("Белки (г):")
        except:
            await update.message.reply_text("Введи число:")

    elif state == "food_protein":
        try:
            context.user_data["food_protein"] = int(text)
            context.user_data["state"] = "food_fat"
            await update.message.reply_text("Жиры (г):")
        except:
            await update.message.reply_text("Введи число:")

    elif state == "food_fat":
        try:
            context.user_data["food_fat"] = int(text)
            context.user_data["state"] = "food_carbs"
            await update.message.reply_text("Углеводы (г):")
        except:
            await update.message.reply_text("Введи число:")

    elif state == "food_carbs":
        try:
            name = context.user_data["food_name"]
            u["food_log"].append({
                "name": name,
                "calories": context.user_data["food_calories"],
                "protein": context.user_data["food_protein"],
                "fat": context.user_data["food_fat"],
                "carbs": int(text),
                "time": datetime.now().strftime("%H:%M"),
            })
            add_xp(u, 10)
            new_ach = check_achievements(u)
            context.user_data.pop("state", None)
            t, m = main_menu(u)
            msg = f"✅ *{name}* добавлено!\n\n{t}"
            if new_ach:
                msg = "🎖 *" + ACHIEVEMENTS[new_ach[0]]["name"] + "!*\n\n" + msg
            await update.message.reply_text(msg, reply_markup=m, parse_mode="Markdown")
        except:
            await update.message.reply_text("Введи число:")

    elif state == "workout_custom_name":
        context.user_data["workout_name"] = text
        context.user_data["state"] = "workout_sets"
        await update.message.reply_text(f"🏋️ *{text}*\n\nПодходов?", parse_mode="Markdown")

    elif state == "workout_sets":
        try:
            context.user_data["workout_sets"] = int(text)
            context.user_data["state"] = "workout_reps"
            await update.message.reply_text("Повторений?")
        except:
            await update.message.reply_text("Введи число:")

    elif state == "workout_reps":
        try:
            context.user_data["workout_reps"] = int(text)
            context.user_data["state"] = "workout_weight"
            await update.message.reply_text("Вес (кг)? Без веса — введи 0")
        except:
            await update.message.reply_text("Введи число:")

    elif state == "workout_weight":
        try:
            weight = float(text.replace(",", "."))
            name = context.user_data["workout_name"]
            sets = context.user_data["workout_sets"]
            reps = context.user_data["workout_reps"]
            u["workout_log"].append({"name": name, "sets": sets, "reps": reps, "weight": weight, "time": datetime.now().strftime("%H:%M")})
            u["total_workouts"] = u.get("total_workouts", 0) + 1
            add_xp(u, 30)

            # Проверка рекорда
            record_text = ""
            records = u.setdefault("records", {})
            if weight > 0:
                prev = records.get(name)
                is_record = not prev or weight > prev["weight"] or (weight == prev["weight"] and reps > prev["reps"])
                if is_record:
                    records[name] = {"weight": weight, "reps": reps, "date": str(date.today())}
                    if prev:
                        record_text = "\n💥 *НОВЫЙ ЛИЧНЫЙ РЕКОРД!* +50 XP"
                        add_xp(u, 50)
                        u["achievements"] = list(set(u.get("achievements", [])) | {"record_broken"})

            new_ach = check_achievements(u)
            context.user_data.pop("state", None)
            motivations = ["Огонь! 🔥", "Так держать! 💪", "Ты машина! 🚀", "Отличная работа! 🏆"]
            t, m = main_menu(u)
            msg = f"✅ *{name}* {sets}×{reps}@{weight}кг! {random.choice(motivations)}{record_text}\n\n{t}"
            if new_ach and new_ach[0] != "record_broken":
                msg = "🎖 *" + ACHIEVEMENTS[new_ach[0]]["name"] + "!*\n\n" + msg
            await update.message.reply_text(msg, reply_markup=m, parse_mode="Markdown")
        except:
            await update.message.reply_text("Введи число:")

    elif state == "set_calorie_goal":
        try:
            u["calorie_goal"] = int(text)
            context.user_data.pop("state", None)
            t, m = main_menu(u)
            await update.message.reply_text(f"✅ Норма: *{u['calorie_goal']} ккал*\n\n{t}", reply_markup=m, parse_mode="Markdown")
        except:
            await update.message.reply_text("Введи число:")

    elif state == "set_water_goal":
        try:
            u["water_goal"] = int(text)
            context.user_data.pop("state", None)
            t, m = main_menu(u)
            await update.message.reply_text(f"✅ Норма воды: *{u['water_goal']} мл*\n\n{t}", reply_markup=m, parse_mode="Markdown")
        except:
            await update.message.reply_text("Введи число:")

    elif state == "update_weight":
        try:
            new_w = float(text.replace(",", "."))
            old_w = u["weight"] or new_w
            u["weight"] = new_w
            u["weight_history"].append({"date": str(date.today()), "weight": new_w})
            diff = new_w - old_w
            sign = "+" if diff > 0 else ""
            if u.get("goal") == "похудеть" and diff <= -5:
                u["achievements"] = list(set(u.get("achievements", [])) | {"weight_lost_5"})
            context.user_data.pop("state", None)
            t, m = main_menu(u)
            await update.message.reply_text(f"⚖️ Вес: *{new_w} кг* ({sign}{diff:.1f} кг)\n\n{t}", reply_markup=m, parse_mode="Markdown")
        except:
            await update.message.reply_text("Введи число:")

    else:
        if u["name"]:
            t, m = main_menu(u)
            await update.message.reply_text(t, reply_markup=m, parse_mode="Markdown")
        else:
            await update.message.reply_text("Напиши /start чтобы начать!")

async def weight_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    u = get_user(uid)
    context.user_data["state"] = "update_weight"
    await update.message.reply_text(f"⚖️ Текущий: *{u['weight']} кг*\n\nВведи новый:", parse_mode="Markdown")

# ── Автоматика ─────────────────────────────────────────────────────────────
async def morning_motivation(context: ContextTypes.DEFAULT_TYPE):
    if datetime.now().hour != 8:
        return
    quote = random.choice(MOTIVATIONAL_QUOTES)
    for uid, u in users.items():
        if u.get("name"):
            try:
                t, m = main_menu(u)
                await context.bot.send_message(uid, f"🌅 *Доброе утро, {u['name']}!*\n\n{quote}\n\n{t}", reply_markup=m, parse_mode="Markdown")
            except:
                pass

async def weekly_report(context: ContextTypes.DEFAULT_TYPE):
    if datetime.now().weekday() != 6 or datetime.now().hour != 20:
        return
    for uid, u in users.items():
        if not u.get("name"):
            continue
        try:
            week_cal = sum(u.get("history", {}).get(str(date.today() - timedelta(days=i)), {}).get("calories", 0) for i in range(7))
            week_w = sum(u.get("history", {}).get(str(date.today() - timedelta(days=i)), {}).get("workouts", 0) for i in range(7))
            week_water = sum(u.get("history", {}).get(str(date.today() - timedelta(days=i)), {}).get("water", 0) for i in range(7))
            text = (
                f"📋 *Недельный отчёт, {u['name']}!*\n\n"
                f"🔥 Средние калории: *{int(week_cal/7)} ккал/день*\n"
                f"🏋️ Тренировок за неделю: *{week_w}*\n"
                f"💧 Средняя вода: *{int(week_water/7)} мл/день*\n"
                f"🔥 Стрик: *{u['streak']} дней*\n"
                f"⭐ XP: *{u.get('xp', 0)}* | {get_level(u.get('xp',0))}\n\n"
                f"{random.choice(MOTIVATIONAL_QUOTES)}"
            )
            chart = make_weekly_chart(u)
            if chart:
                await context.bot.send_photo(uid, photo=chart, caption=text, parse_mode="Markdown")
            else:
                await context.bot.send_message(uid, text, parse_mode="Markdown")
        except:
            pass

async def send_reminders(context: ContextTypes.DEFAULT_TYPE):
    hour = datetime.now().hour
    for uid, u in users.items():
        try:
            if u["reminders"]["water"] and u["water"] < u["water_goal"] and hour % 2 == 0:
                await context.bot.send_message(uid, f"💧 Не забывай пить! {u['water']}мл из {u['water_goal']}мл")
            if u["reminders"]["food"] and 8 <= hour <= 21:
                total_cal, _, _, _ = calories_summary(u)
                if total_cal < u["calorie_goal"] * 0.5 and hour in (12, 18):
                    await context.bot.send_message(uid, f"🍽 Время поесть! {total_cal}/{u['calorie_goal']} ккал")
        except:
            pass

# ── Запуск ─────────────────────────────────────────────────────────────────
def main():
    app = Application.builder().token(BOT_TOKEN).build()
    conv = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            SETUP_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, setup_name)],
            SETUP_WEIGHT: [MessageHandler(filters.TEXT & ~filters.COMMAND, setup_weight)],
            SETUP_HEIGHT: [MessageHandler(filters.TEXT & ~filters.COMMAND, setup_height)],
            SETUP_AGE: [MessageHandler(filters.TEXT & ~filters.COMMAND, setup_age)],
            SETUP_GOAL: [CallbackQueryHandler(setup_goal, pattern="^goal_")],
        },
        fallbacks=[CommandHandler("start", start)],
    )
    app.add_handler(conv)
    app.add_handler(CommandHandler("menu", menu))
    app.add_handler(CommandHandler("weight", weight_command))
    app.add_handler(CallbackQueryHandler(button_handler))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, text_handler))
    app.job_queue.run_repeating(send_reminders, interval=3600, first=60)
    app.job_queue.run_repeating(morning_motivation, interval=3600, first=10)
    app.job_queue.run_repeating(weekly_report, interval=3600, first=30)
    print("💪 FitBot Pro запущен!")
    app.run_polling()

if __name__ == "__main__":
    main()
